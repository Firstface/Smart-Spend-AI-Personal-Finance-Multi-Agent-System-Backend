"""
File parsers: WeChat Pay Excel + Alipay CSV.

WeChat bill structure:
  First 17 rows are header notes, row 18 is column headers, data starts at row 19.
  Columns: transaction_time | type | counterparty | goods | in/out | amount(CNY) | payment | status | order_id | merchant_order_id | remark

Alipay bill structure:
  First several rows are header notes; find the row starting with "交易时间," as column header.
  Encoding: GBK.
  Columns: 交易时间 | 交易分类 | 交易对方 | 对方账号 | 商品说明 | 收/支 | 金额 | 收/付款方式 | 交易状态 | 交易订单号 | 商家订单号 | 备注
"""
import openpyxl
import csv
import io
import logging
from datetime import datetime
from typing import List

from schemas.transaction import TransactionRaw, DirectionEnum

logger = logging.getLogger("categorization.parser")


# ── WeChat Pay Excel ────────────────────────────────────────────────────────────
def parse_wechat_excel(content: bytes) -> List[TransactionRaw]:
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    transactions = []

    direction_map = {
        "支出": DirectionEnum.EXPENSE,
        "收入": DirectionEnum.INCOME,
    }

    for row in ws.iter_rows(min_row=19, values_only=True):
        if row[0] is None:
            break

        txn_type = str(row[1] or "").strip()
        direction_raw = str(row[4] or "").strip()
        direction = direction_map.get(direction_raw, DirectionEnum.NEUTRAL)

        # Refund transactions are marked as neutral
        if "退款" in txn_type:
            direction = DirectionEnum.NEUTRAL

        # Parse amount: strip possible ¥ symbol
        raw_amount = str(row[5] or "0").replace("¥", "").replace(",", "").strip()
        try:
            amount = abs(float(raw_amount))
        except ValueError:
            logger.warning(f"WeChat bill amount parse failed: {row[5]}, skipping row")
            continue

        # Parse datetime
        if isinstance(row[0], datetime):
            txn_time = row[0]
        else:
            try:
                txn_time = datetime.strptime(str(row[0]).strip(), "%Y-%m-%d %H:%M:%S")
            except ValueError:
                logger.warning(f"WeChat bill datetime parse failed: {row[0]}, skipping row")
                continue

        txn = TransactionRaw(
            source="wechat",
            transaction_time=txn_time,
            transaction_type=txn_type,
            counterparty=str(row[2] or "").strip(),
            goods_description=str(row[3] or "").strip() or None,
            direction=direction,
            amount=amount,
            payment_method=str(row[6] or "").strip() or None,
            status=str(row[7] or "").strip() or None,
            order_id=str(row[8] or "").strip() or None,
            merchant_order_id=str(row[9] or "").strip() or None,
            remark=str(row[10] or "").strip() if row[10] and str(row[10]).strip() != "/" else None,
            original_category=None,  # WeChat has no built-in category
        )
        transactions.append(txn)

    logger.info(f"WeChat bill parsed: {len(transactions)} records")
    return transactions


# ── Datetime helper ─────────────────────────────────────────────────────────────
_DATETIME_FORMATS = [
    "%Y-%m-%d %H:%M:%S",   # 2025-12-17 12:17:00  (standard Alipay)
    "%Y-%m-%d %H:%M",      # 2025-12-17 12:17
    "%d/%m/%Y %H:%M:%S",   # 17/12/2025 12:17:00  (some exports)
    "%d/%m/%Y %H:%M",      # 17/12/2025 12:17
    "%m/%d/%Y %H:%M:%S",   # 12/17/2025 12:17:00
    "%m/%d/%Y %H:%M",      # 12/17/2025 12:17
    "%Y/%m/%d %H:%M:%S",   # 2025/12/17 12:17:00
    "%Y/%m/%d %H:%M",      # 2025/12/17 12:17
]

def _parse_datetime(value: str) -> datetime | None:
    """Try multiple datetime formats; return None if all fail."""
    value = value.strip()
    for fmt in _DATETIME_FORMATS:
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    return None


# ── Alipay CSV ──────────────────────────────────────────────────────────────────
def parse_alipay_csv(content: bytes) -> List[TransactionRaw]:
    # Alipay bills are GBK encoded
    try:
        text = content.decode("gbk")
    except UnicodeDecodeError:
        text = content.decode("utf-8", errors="replace")

    lines = text.strip().split("\n")

    # Find the column header row (starts with "交易时间,")
    header_idx = None
    for i, line in enumerate(lines):
        if line.startswith("交易时间,") or line.startswith("交易时间，"):
            header_idx = i
            break
    if header_idx is None:
        raise ValueError("Cannot identify Alipay CSV format: column header row not found (should start with '交易时间,')")

    # Filter trailing summary rows (Alipay CSV ends with "---" separated totals)
    data_lines = []
    for line in lines[header_idx:]:
        if line.startswith("---") or line.strip() == "":
            break
        data_lines.append(line)

    reader = csv.DictReader(io.StringIO("\n".join(data_lines)))
    transactions = []

    direction_map = {
        "支出": DirectionEnum.EXPENSE,
        "收入": DirectionEnum.INCOME,
        "不计收支": DirectionEnum.NEUTRAL,
    }

    for row in reader:
        time_str = row.get("交易时间", "").strip()
        if not time_str:
            continue

        txn_time = _parse_datetime(time_str)
        if txn_time is None:
            logger.warning(f"Alipay bill datetime parse failed: {time_str}, skipping row")
            continue

        direction_str = row.get("收/支", "").strip()
        direction = direction_map.get(direction_str, DirectionEnum.NEUTRAL)

        category_raw = row.get("交易分类", "").strip()
        if category_raw == "退款":
            direction = DirectionEnum.NEUTRAL

        raw_amount = row.get("金额", "0").strip().replace(",", "")
        try:
            amount = abs(float(raw_amount))
        except ValueError:
            logger.warning(f"Alipay bill amount parse failed: {row.get('金额')}, skipping row")
            continue

        txn = TransactionRaw(
            source="alipay",
            transaction_time=txn_time,
            transaction_type=category_raw or None,
            counterparty=row.get("交易对方", "").strip(),
            counterparty_account=row.get("对方账号", "").strip() or None,
            goods_description=row.get("商品说明", "").strip() or None,
            direction=direction,
            amount=amount,
            payment_method=row.get("收/付款方式", "").strip() or None,
            status=row.get("交易状态", "").strip() or None,
            order_id=row.get("交易订单号", "").strip() or None,
            merchant_order_id=row.get("商家订单号", "").strip() or None,
            original_category=category_raw if category_raw and category_raw != "退款" else None,
            remark=row.get("备注", "").strip() or None,
        )
        transactions.append(txn)

    logger.info(f"Alipay bill parsed: {len(transactions)} records")
    return transactions


# ── Alipay Excel ────────────────────────────────────────────────────────────────
def parse_alipay_excel(content: bytes) -> List[TransactionRaw]:
    """
    Parse Alipay bill exported as .xlsx.
    Column layout mirrors the CSV version; find the header row by looking for '交易时间'.
    """
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active

    # Find the header row
    header_row_idx = None
    headers = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row and str(row[0] or "").strip() == "交易时间":
            header_row_idx = i
            headers = [str(c or "").strip() for c in row]
            break

    if header_row_idx is None:
        raise ValueError("Cannot identify Alipay Excel format: header row with '交易时间' not found")

    direction_map = {
        "支出": DirectionEnum.EXPENSE,
        "收入": DirectionEnum.INCOME,
        "不计收支": DirectionEnum.NEUTRAL,
    }

    transactions = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not row or row[0] is None or str(row[0]).startswith("---"):
            break
        data = {headers[i]: str(v or "").strip() for i, v in enumerate(row) if i < len(headers)}

        time_str = data.get("交易时间", "")
        if not time_str:
            continue
        txn_time = _parse_datetime(time_str)
        if txn_time is None:
            logger.warning(f"Alipay Excel datetime parse failed: {time_str}, skipping row")
            continue

        direction_str = data.get("收/支", "").strip()
        direction = direction_map.get(direction_str, DirectionEnum.NEUTRAL)

        category_raw = data.get("交易分类", "").strip()
        if category_raw == "退款":
            direction = DirectionEnum.NEUTRAL

        raw_amount = data.get("金额", "0").replace(",", "")
        try:
            amount = abs(float(raw_amount))
        except ValueError:
            logger.warning(f"Alipay Excel amount parse failed: {data.get('金额')}, skipping row")
            continue

        txn = TransactionRaw(
            source="alipay",
            transaction_time=txn_time,
            transaction_type=category_raw or None,
            counterparty=data.get("交易对方", ""),
            counterparty_account=data.get("对方账号", "") or None,
            goods_description=data.get("商品说明", "") or None,
            direction=direction,
            amount=amount,
            payment_method=data.get("收/付款方式", "") or None,
            status=data.get("交易状态", "") or None,
            order_id=data.get("交易订单号", "") or None,
            merchant_order_id=data.get("商家订单号", "") or None,
            original_category=category_raw if category_raw and category_raw != "退款" else None,
            remark=data.get("备注", "") or None,
        )
        transactions.append(txn)

    logger.info(f"Alipay Excel parsed: {len(transactions)} records")
    return transactions


# ── Unified entry point ─────────────────────────────────────────────────────────
def parse_file(filename: str, content: bytes) -> List[TransactionRaw]:
    """
    Dispatch to the appropriate parser based on file extension and content.
    - .csv               → Alipay CSV parser
    - .xlsx/.xls         → Try WeChat parser first; if zero records, try Alipay Excel parser
    """
    name_lower = filename.lower()
    if name_lower.endswith(".csv"):
        return parse_alipay_csv(content)
    elif name_lower.endswith(".xlsx") or name_lower.endswith(".xls"):
        # Try WeChat first (standard .xlsx format)
        try:
            result = parse_wechat_excel(content)
            if result:
                return result
        except Exception:
            pass
        # Fallback: try Alipay Excel format
        logger.info(f"WeChat parse yielded 0 records for '{filename}', retrying as Alipay Excel")
        return parse_alipay_excel(content)
    else:
        raise ValueError(f"Unsupported file format: {filename}. Please upload .xlsx (WeChat/Alipay) or .csv (Alipay).")
